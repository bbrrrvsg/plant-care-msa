"""PlantCare MSA API 명세서 (실무 양식)"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "API 명세서"

# ============ 스타일 ============
SECTION_FILL = PatternFill("solid", start_color="2D5A27")
HEADER_FILL = PatternFill("solid", start_color="3A7D44")
ALT_FILL = PatternFill("solid", start_color="F5FAF5")
BODY_FONT = Font(name="맑은 고딕", size=10)
HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
SECTION_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=12)
MONO_FONT = Font(name="Consolas", size=9)

THIN = Side(border_style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

HEADERS = [
    "no", "기능명(한글)", "기능(영문)", "기능설명", "HTTP 메소드", "HTTP URL",
    "매개변수\nContent-Type", "매개변수 타입", "매개변수 샘플",
    "반환타입", "반환 샘플(back)", "담당자", "작업여부", "비고"
]

# 섹션별 데이터
# (한글명, 영문명, 설명, METHOD, URL, ContentType, 타입, 매개변수샘플, 반환타입, 반환샘플, 비고)
SECTIONS = [
    ("회원 / 인증 관리", [
        ("회원가입", "signUp", "신규 사용자 회원가입", "POST", "/auth/signup",
         "application/json", "SignUpRequestDto",
         '{"userId":"plant01","password":"pw1234!","nickname":"플랜터","email":"abc@a.com"}',
         "ApiResponse", '{"success":true,"data":"회원가입 성공"}', ""),
        ("로그인", "login", "JWT 토큰 발급", "POST", "/auth/login",
         "application/json", "LoginRequestDto",
         '{"userId":"plant01","password":"pw1234!"}',
         "ApiResponse<TokenResponseDto>",
         '{"success":true,"data":{"accessToken":"eyJhbGc...","userId":"plant01"}}', ""),
        ("사용자 조회", "getUser", "사용자 정보 단건 조회", "GET", "/auth/user/{userId}",
         "-", "PathVariable", "userId : plant01",
         "UserResponseDto",
         '{"userId":"plant01","nickname":"플랜터","email":"abc@a.com"}', "JWT 필요"),
    ]),
    ("내 식물 관리", [
        ("내 식물 전체 조회", "getMyPlants", "사용자가 등록한 식물 목록", "GET", "/plant",
         "-", "QueryParam", "userId : plant01",
         "List<PlantResponseDto>",
         '[{"myPlantId":1,"nickname":"우리집몬스","plantName":"몬스테라"}]', ""),
        ("내 식물 단건 조회", "getMyPlant", "식물 상세 정보 조회", "GET", "/plant/{myPlantId}",
         "-", "PathVariable", "myPlantId : 1",
         "PlantResponseDto",
         '{"myPlantId":1,"nickname":"우리집몬스","imageUrl":"..."}', ""),
        ("식물 ID 목록", "getPlantIdsByUserId", "사용자의 식물 ID 리스트", "GET", "/plant/ids",
         "-", "QueryParam", "userId : plant01",
         "List<Integer>", "[1, 2, 3]", ""),
        ("내 식물 등록", "addMyPlant", "식물 + 이미지 업로드", "POST", "/plant",
         "multipart/form-data", "FormData",
         "userId, nickname, location, deviceId, deviceName, image(file)",
         "PlantResponseDto",
         '{"myPlantId":4,"nickname":"새식물","imageUrl":"..."}', "이미지 업로드"),
        ("내 식물 수정", "updateMyPlant", "식물 정보 수정", "PUT", "/plant/{myPlantId}",
         "application/json", "PlantRequestDto",
         '{"nickname":"새이름","location":"거실"}',
         "PlantResponseDto", '{"myPlantId":1,"nickname":"새이름"}', ""),
        ("내 식물 삭제", "deleteMyPlant", "식물 삭제", "DELETE", "/plant/{myPlantId}",
         "-", "PathVariable", "myPlantId : 1",
         "void", "204 No Content", ""),
        ("센서 데이터 조회", "getSensorDataByPlantId", "식물의 최신 센서값 (sensor-service 연동)",
         "GET", "/plant/sensor/{plantId}",
         "-", "PathVariable", "plantId : 1",
         "SensorDataDto",
         '{"plantId":1,"temperature":24.5,"soilMoisture":42}', "FeignClient"),
        ("식물 AI 진단", "diagnosePlant", "식물 진단 요청 (ai-service 연동)",
         "POST", "/plant/{myPlantId}/diagnosis",
         "multipart/form-data", "FormData",
         "myPlantId(path), image(file)",
         "AIDiagnosisDto",
         '{"diagnosisId":12,"title":"잎마름병","subtitle":"..."}', "FeignClient"),
    ]),
    ("식물 도감", [
        ("도감 전체 조회", "getAllBooks", "식물 도감 리스트", "GET", "/book",
         "-", "-", "-",
         "List<BookDto>",
         '[{"speciesCode":101,"plantName":"몬스테라"}]', ""),
        ("도감 단건 조회", "getBook", "식물 도감 상세", "GET", "/book/{speciesCode}",
         "-", "PathVariable", "speciesCode : 101",
         "BookDto",
         '{"speciesCode":101,"plantName":"몬스테라","careInfo":"..."}', ""),
        ("식물 이름 검색", "searchBooks", "식물명으로 도감 검색", "GET", "/book/search",
         "-", "QueryParam", "name : 몬스테라",
         "List<BookDto>",
         '[{"speciesCode":101,"plantName":"몬스테라"}]', ""),
        ("농사로 API 적재", "fetchAndSave", "농사로 Open API 데이터 DB 저장",
         "POST", "/book/fetch",
         "-", "-", "-",
         "String", '"저장완료"', "최초 1회 / 관리자"),
    ]),
    ("성장 일지", [
        ("성장 일지 목록", "getList", "사용자의 일지 전체 조회", "GET", "/growth-log",
         "-", "QueryParam", "userId : 1",
         "List<GrowthLogDto>",
         '[{"logId":1,"title":"오늘의 일지","logDate":"2026-05-20"}]', ""),
        ("성장 일지 상세", "getDetailLog", "일지 상세 조회 (+진단 포함 가능)",
         "GET", "/growth-log/{logId}",
         "-", "PathVariable + QueryParam",
         "logId : 1, includeDiagnosis : false",
         "GrowthLogDto",
         '{"logId":1,"title":"...","content":"...","photoUrl":"..."}', ""),
        ("성장 일지 작성", "create", "일지 신규 등록", "POST", "/growth-log/write",
         "application/json", "GrowthLogRequestDto",
         '{"plantId":1,"title":"잘 자라요","content":"...","type":"성장 기록"}',
         "GrowthLogDto",
         '{"logId":5,"title":"잘 자라요"}', ""),
        ("성장 일지 수정", "update", "일지 내용 수정", "PUT", "/growth-log/{logId}",
         "application/json", "GrowthLogDto",
         '{"title":"수정됨","content":"..."}',
         "GrowthLogDto", '{"logId":1,"title":"수정됨"}', ""),
        ("성장 일지 삭제", "delete", "일지 삭제", "DELETE", "/growth-log/{logId}",
         "-", "PathVariable", "logId : 1",
         "boolean", "true", ""),
    ]),
    ("센서 디바이스 (IoT)", [
        ("디바이스 자동 등록", "registerDevice", "ESP32 전원 ON 시 자동 등록",
         "POST", "/api/sensor/device/register",
         "application/json", "Map<String,String>",
         '{"deviceId":"ESP32-AB12"}',
         "void", "200 OK", "ESP32에서 호출"),
        ("디바이스 별명 설정", "updateDeviceName", "기기 별명 변경",
         "PATCH", "/api/sensor/device/{deviceId}/name",
         "application/json", "Map<String,String>",
         '{"deviceName":"베란다 센서"}',
         "void", "200 OK", ""),
        ("디바이스-식물 연결", "linkPlant", "디바이스에 식물 연결",
         "PATCH", "/api/sensor/device/{deviceId}/link",
         "application/json", "SensorDeviceDto",
         '{"plantId":1,"userId":"plant01","deviceName":"베란다","speciesCode":101}',
         "void", "200 OK", ""),
        ("디바이스-식물 해제", "unlinkPlant", "디바이스에서 식물 연결 해제",
         "PATCH", "/api/sensor/device/{deviceId}/unlink",
         "-", "PathVariable", "deviceId : ESP32-AB12",
         "void", "200 OK", ""),
        ("디바이스 비활성화", "deactivateDevice", "기기 꺼짐 감지 시",
         "PATCH", "/api/sensor/device/{deviceId}/deactivate",
         "-", "PathVariable", "deviceId : ESP32-AB12",
         "void", "200 OK", ""),
        ("디바이스 상세 조회", "detailDevice", "디바이스 상세 정보",
         "GET", "/api/sensor/device/{deviceId}",
         "-", "PathVariable", "deviceId : ESP32-AB12",
         "SensorDeviceDto",
         '{"deviceId":"ESP32-AB12","deviceName":"베란다","plantId":1}', ""),
        ("미연결 디바이스 목록", "getUnlinkedDevices", "식물 연결 안 된 디바이스 목록",
         "GET", "/api/sensor/device/unlinked",
         "-", "-", "-",
         "List<SensorDeviceDto>",
         '[{"deviceId":"ESP32-CD34","deviceName":null}]', ""),
        ("내 디바이스 목록", "getMyDevices", "사용자의 활성 디바이스 전체",
         "GET", "/api/sensor/device",
         "-", "QueryParam", "userId : plant01",
         "List<SensorDeviceDto>",
         '[{"deviceId":"ESP32-AB12","plantId":1}]', ""),
    ]),
    ("센서 데이터", [
        ("센서 데이터 수신", "receiveData", "ESP32 → Redis 저장",
         "POST", "/api/sensor/data",
         "application/json", "SensorDataDto",
         '{"plantId":1,"temperature":24.5,"humidity":60,"soilMoisture":42}',
         "void", "200 OK", "ESP32 주기 호출"),
        ("최신 센서값 조회", "getLatestData", "Redis에서 최신 데이터 조회",
         "GET", "/api/sensor/data/{plantId}",
         "-", "PathVariable", "plantId : 1",
         "SensorDataDto",
         '{"plantId":1,"temperature":24.5,"soilMoisture":42}', "없으면 204"),
        ("시간별 평균 히스토리", "getHistory", "대시보드 차트용 히스토리",
         "GET", "/api/sensor/data/{plantId}/history",
         "-", "PathVariable + QueryParam",
         "plantId : 1, hours : 10",
         "List<SensorDataDto>",
         '[{"hour":"2026-05-20T10","temperature":24}]', ""),
    ]),
    ("AI 진단 (Gemini)", [
        ("AI 진단 (Gemini)", "diagnosePlant", "Gemini 기반 식물 진단",
         "POST", "/ai/gemini",
         "multipart/form-data", "FormData",
         "image(file), plantId(optional)",
         "AIDiagnosisDto",
         '{"diagnosisId":1,"title":"잎마름병","subtitle":"수분부족"}', "Gemini API"),
        ("식물 종 식별", "identifyPlant", "사진으로 식물 종 식별",
         "POST", "/ai/identify",
         "multipart/form-data", "FormData", "image(file)",
         "List<BookDto>",
         '[{"speciesCode":101,"plantName":"몬스테라"}]', "Gemini API"),
        ("진단 목록 조회", "User_AIList", "사용자의 진단 기록 전체",
         "GET", "/diagnosis/Ai",
         "-", "QueryParam", "userId : 1",
         "List<Map>",
         '[{"diagnosisId":1,"title":"잎마름병","diagnosisDate":"2026-05-20"}]', ""),
        ("진단 상세 조회", "User_Details", "진단 결과 상세",
         "GET", "/diagnosis/{diagnosisId}",
         "-", "PathVariable", "diagnosisId : 1",
         "AIDiagnosisDto",
         '{"diagnosisId":1,"title":"잎마름병","details":"..."}', ""),
        ("재진단 요청", "reDiagnose", "기존 진단 ID로 재진단",
         "PUT", "/diagnosis/{diagnosisId}",
         "multipart/form-data", "FormData", "image(file)",
         "AIDiagnosisDto",
         '{"diagnosisId":1,"title":"재진단결과"}', "Gemini API"),
        ("진단 삭제", "delete", "진단 기록 삭제",
         "DELETE", "/diagnosis/delete",
         "-", "QueryParam", "diagnosisId : 1",
         "boolean", "true", ""),
    ]),
]

# ============ 작성 ============
ws.sheet_view.showGridLines = False
row = 1

# 컬럼 너비
widths = [5, 18, 22, 28, 11, 32, 16, 22, 36, 22, 38, 10, 9, 16]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = w

global_no = 1

for section_name, items in SECTIONS:
    # 섹션 헤더
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    sc = ws.cell(row=row, column=1, value=f"  ▌ {section_name}")
    sc.font = SECTION_FONT
    sc.fill = SECTION_FILL
    sc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 26
    row += 1

    # 컬럼 헤더
    for i, h in enumerate(HEADERS, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 32
    row += 1

    # 데이터
    for idx, item in enumerate(items):
        kor, eng, desc, method, url, ctype, ptype, psample, rtype, rsample, note = item
        values = [global_no, kor, eng, desc, method, url, ctype, ptype, psample,
                  rtype, rsample, "이한승", "완료", note]

        alt = idx % 2 == 1
        for ci, val in enumerate(values, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.border = BORDER
            c.font = BODY_FONT
            c.alignment = LEFT
            if alt:
                c.fill = ALT_FILL

            # 특수 컬럼 처리
            if ci == 1:  # no
                c.alignment = CENTER
            elif ci == 3:  # 영문명
                c.font = Font(name="Consolas", size=9, color="2D5A27", bold=True)
            elif ci == 5:  # METHOD
                c.alignment = CENTER
                c.font = Font(name="Consolas", bold=True, size=10, color="FFFFFF")
                method_colors = {"GET": "61AFFE", "POST": "49CC90", "PUT": "FCA130",
                                 "PATCH": "50E3C2", "DELETE": "F93E3E"}
                c.fill = PatternFill("solid",
                                     start_color=method_colors.get(method, "808080"))
            elif ci == 6:  # URL
                c.font = Font(name="Consolas", size=9, color="2D5A27", bold=True)
            elif ci in (9, 11):  # 매개변수 샘플, 반환 샘플
                c.font = Font(name="Consolas", size=8, color="374151")
            elif ci in (7, 13):  # ContentType, 작업여부
                c.alignment = CENTER
            elif ci == 12:  # 담당자
                c.alignment = CENTER

        ws.row_dimensions[row].height = 40
        row += 1
        global_no += 1

    row += 1  # 섹션 사이 빈 줄

# 헤더 고정 (첫 섹션 헤더 다음)
ws.freeze_panes = "A3"

output = r"C:\Users\korel\Development\plant-care-msa\docs\PlantCare_API_명세서.xlsx"
wb.save(output)
print(f"저장 완료: {output}")
print(f"총 API 수: {global_no - 1}")
